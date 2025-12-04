import io
import re
import unicodedata
from difflib import get_close_matches
from typing import Optional

from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
import os
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-me-to-a-secure-key")

# Canonical columns to check
COLUMNS_TO_CHECK = ['Registration No.', 'Student Name', 'Contact No.', 'Email Address']

# Colors
ROW_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for matched rows
CELL_ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for trigger cells

# ------------------------
# Helpers: normalization & header detection
# ------------------------
def normalize_col_name(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ").replace("\u200B", "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch)[0] != "C")
    s = " ".join(s.split())
    s = "".join(ch for ch in s if ch.isalnum() or ch.isspace())
    return s.lower()


def detect_header_row_and_read(excel_fileobj=None, url: Optional[str] = None, max_header_scan=10) -> pd.DataFrame:
    try:
        if excel_fileobj:
            try:
                excel_fileobj.seek(0)
            except Exception:
                pass
            preview = pd.read_excel(excel_fileobj, header=None, nrows=max_header_scan, engine="openpyxl")
            try:
                excel_fileobj.seek(0)
            except Exception:
                pass
            source_bytes = None
        else:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            content = r.content
            preview = pd.read_excel(io.BytesIO(content), header=None, nrows=max_header_scan, engine="openpyxl")
            source_bytes = content
    except Exception:
        if excel_fileobj:
            try:
                excel_fileobj.seek(0)
            except Exception:
                pass
            return pd.read_excel(excel_fileobj, engine="openpyxl")
        else:
            return pd.read_excel(io.BytesIO(source_bytes), engine="openpyxl")

    header_row_idx = None
    for idx in range(min(len(preview), max_header_scan)):
        row = preview.iloc[idx].tolist()
        non_null_count = sum(1 for v in row if not (pd.isna(v) or str(v).strip() == ""))
        str_like_count = sum(1 for v in row if isinstance(v, str) and v.strip() != "")
        if non_null_count >= max(1, len(row) // 2) and str_like_count >= max(1, len(row) // 3):
            header_row_idx = idx
            break

    try:
        if header_row_idx is not None:
            if excel_fileobj:
                try:
                    excel_fileobj.seek(0)
                except Exception:
                    pass
                df = pd.read_excel(excel_fileobj, header=header_row_idx, engine="openpyxl")
            else:
                df = pd.read_excel(io.BytesIO(source_bytes), header=header_row_idx, engine="openpyxl")
            return df
    except Exception:
        pass

    if excel_fileobj:
        try:
            excel_fileobj.seek(0)
        except Exception:
            pass
        return pd.read_excel(excel_fileobj, engine="openpyxl")
    else:
        return pd.read_excel(io.BytesIO(source_bytes), engine="openpyxl")


# ------------------------
# Read wrapper with diagnostics
# ------------------------
def read_excel_with_diagnostics(fileobj, url: Optional[str], label: str) -> pd.DataFrame:
    if fileobj:
        try:
            return detect_header_row_and_read(excel_fileobj=fileobj)
        except Exception as e:
            raise RuntimeError(f"{label}: Failed to read uploaded file — {e}")

    if not url:
        raise RuntimeError(f"{label}: No file or URL provided")

    try:
        r = requests.get(url, timeout=30)
    except Exception as e:
        raise RuntimeError(f"{label}: Failed to fetch URL ({url}) — {e}")

    if r.status_code != 200:
        raise RuntimeError(f"{label}: URL returned status code {r.status_code}. Make sure link is direct-download and public.")

    content_type = (r.headers.get("Content-Type") or "").lower()
    if "html" in content_type or (r.text and r.text.strip().lower().startswith("<html")):
        snippet = (r.text[:400] + "...") if r.text else ""
        raise RuntimeError(
            f"{label}: The URL returned HTML (likely a login page or an error page) instead of an Excel file.\n"
            f"Preview of server response (first 400 chars):\n{repr(snippet)}\n"
            f"Tip: if this is a SharePoint/OneDrive link, either create a direct-download link or download and upload the file manually."
        )
    try:
        return detect_header_row_and_read(url=url)
    except Exception as e:
        raise RuntimeError(f"{label}: Downloaded content could not be read as Excel — {e}. Ensure link points to a valid .xlsx file.")


# ------------------------
# Normalization for values (matching)
# ------------------------
def normalize_value(col: str, val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s == "":
        return None
    if col == 'Contact No.':
        digits = re.sub(r'\D', '', s)
        return digits if digits else None
    if col == 'Email Address':
        return s.lower()
    if col == 'Student Name':
        s = re.sub(r'\s+', ' ', s)
        return s.lower()
    if col == 'Registration No.':
        return s.upper()
    return s.lower()


def build_index(df: pd.DataFrame, cols):
    idx = {}
    for col in cols:
        if col in df.columns:
            idx[col] = {}
            for i, raw in enumerate(df[col].tolist()):
                norm = normalize_value(col, raw)
                if norm is None:
                    continue
                idx[col].setdefault(norm, []).append(i)
    return idx


def find_matches(df1: pd.DataFrame, df2: pd.DataFrame, cols):
    idx1 = build_index(df1, cols)
    idx2 = build_index(df2, cols)
    matches1 = {}
    matches2 = {}
    for col in cols:
        if col not in idx1 or col not in idx2:
            continue
        for val, rows1 in idx1[col].items():
            rows2 = idx2[col].get(val)
            if not rows2:
                continue
            for r1 in rows1:
                rec1 = matches1.setdefault(r1, {"matched_cols": set(), "matched_with": {}})
                rec1["matched_cols"].add(col)
                rec1["matched_with"].setdefault(col, []).extend(rows2)
            for r2 in rows2:
                rec2 = matches2.setdefault(r2, {"matched_cols": set(), "matched_with": {}})
                rec2["matched_cols"].add(col)
                rec2["matched_with"].setdefault(col, []).extend(rows1)
    for d in (matches1, matches2):
        for k, v in d.items():
            v["matched_cols"] = sorted(list(v["matched_cols"]))
            for c, lst in v["matched_with"].items():
                v["matched_with"][c] = sorted(list(set(lst)))
    return matches1, matches2


# ------------------------
# Workbook builders: produce one sheet per source (sheet1 & sheet2)
# ------------------------
def build_result_sheet(ws, df: pd.DataFrame, matches: dict):
    """
    Given an openpyxl worksheet ws and df+matches for that df:
      - write header
      - write non-matched rows first then matched rows appended
      - color matched rows yellow and trigger cells orange
      - NO extra columns added
    """
    headers = list(df.columns)
    ws.append(headers)

    matched_indices = sorted(matches.keys())
    matched_set = set(matched_indices)
    all_indices = list(range(len(df)))
    non_matched_indices = [i for i in all_indices if i not in matched_set]
    ordered_indices = non_matched_indices + matched_indices

    idx_to_rownum = {}
    for out_r, orig_idx in enumerate(ordered_indices, start=2):
        idx_to_rownum[orig_idx] = out_r
        row_vals = [df.at[orig_idx, c] if c in df.columns else "" for c in df.columns]
        ws.append(row_vals)

    # color matched rows yellow
    for orig_idx in matched_indices:
        rownum = idx_to_rownum[orig_idx]
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=rownum, column=col_idx).fill = ROW_YELLOW

    # color specific trigger cells orange (override)
    col_name_to_idx = {c: i + 1 for i, c in enumerate(list(df.columns))}
    for orig_idx, info in matches.items():
        rownum = idx_to_rownum[orig_idx]
        for matched_col in info["matched_cols"]:
            if matched_col in col_name_to_idx:
                col_idx = col_name_to_idx[matched_col]
                ws.cell(row=rownum, column=col_idx).fill = CELL_ORANGE

    # header bold
    try:
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
    except Exception:
        pass

    # autosize
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = cell.value
                l = len(str(v)) if v is not None else 0
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)


# ------------------------
# Flask routes
# ------------------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        f1 = request.files.get("file1")
        f2 = request.files.get("file2")
        url1 = (request.form.get("url1") or "").strip()
        url2 = (request.form.get("url2") or "").strip()

        if not (f1 or url1) or not (f2 or url2):
            flash("Please provide both sheet inputs (file upload or public direct URL).", "danger")
            return redirect(url_for("index"))

        try:
            df1 = read_excel_with_diagnostics(f1.stream if f1 else None, url1 if (not f1) else None, label="Sheet 1")
        except RuntimeError as e:
            flash(str(e), "danger")
            return redirect(url_for("index"))

        try:
            df2 = read_excel_with_diagnostics(f2.stream if f2 else None, url2 if (not f2) else None, label="Sheet 2")
        except RuntimeError as e:
            flash(str(e), "danger")
            return redirect(url_for("index"))

        # detect normalized columns present in both and give diagnostic if none
        orig_cols_1 = list(df1.columns)
        orig_cols_2 = list(df2.columns)
        norm_map_1 = {c: normalize_col_name(c) for c in orig_cols_1}
        norm_map_2 = {c: normalize_col_name(c) for c in orig_cols_2}
        norm_set_1 = set(norm_map_1.values())
        norm_set_2 = set(norm_map_2.values())
        wanted_norm = {c: normalize_col_name(c) for c in COLUMNS_TO_CHECK}

        intersect_norm = norm_set_1.intersection(norm_set_2).intersection(set(wanted_norm.values()))
        cols_to_compare = []
        for orig_wanted, norm_w in wanted_norm.items():
            if norm_w in intersect_norm:
                cols_to_compare.append(orig_wanted)

        if not cols_to_compare:
            def pretty_list(cols):
                return "\n".join([f"  - original: {repr(c)}   normalized: {repr(normalize_col_name(c))}" for c in cols])

            msg = (
                "No common comparison columns found in both sheets. Diagnostics below:\n\n"
                "Sheet 1 columns:\n" + pretty_list(orig_cols_1) + "\n\n"
                "Sheet 2 columns:\n" + pretty_list(orig_cols_2) + "\n\n"
            )
            suggestions = []
            for n1 in norm_set_1:
                close = get_close_matches(n1, list(wanted_norm.values()), n=2, cutoff=0.7)
                if close:
                    suggestions.append(f"Sheet1 normalized '{n1}' ~ matches wanted {close}")
            for n2 in norm_set_2:
                close = get_close_matches(n2, list(wanted_norm.values()), n=2, cutoff=0.7)
                if close:
                    suggestions.append(f"Sheet2 normalized '{n2}' ~ matches wanted {close}")

            if suggestions:
                msg += "Possible fuzzy suggestions:\n" + "\n".join("  - " + s for s in suggestions) + "\n\n"

            msg += "Tips: retype headers in Excel to remove hidden characters, remove trailing punctuation, or upload files after saving as fresh .xlsx."
            flash(msg, "danger")
            return redirect(url_for("index"))

        # compute matches
        matches1, matches2 = find_matches(df1, df2, COLUMNS_TO_CHECK)

        # build workbook with two sheets
        wb = Workbook()
        # sheet for df1
        ws1 = wb.active
        ws1.title = "Result_Sheet1"
        build_result_sheet(ws1, df1, matches1)

        # sheet for df2 (create new)
        ws2 = wb.create_sheet(title="Result_Sheet2")
        build_result_sheet(ws2, df2, matches2)

        # send file
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        return send_file(fp,
                         as_attachment=True,
                         download_name="repeated_entries.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template("index.html", cols=COLUMNS_TO_CHECK)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
